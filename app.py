import os
import uuid
import datetime
import shutil
from copy import copy
from typing import Dict, List, Optional
from pydantic import BaseModel

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from openpyxl import load_workbook

app = FastAPI(title="FlexiMapper API")

# Allow CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_headers=["*"],
    allow_methods=["*"],
)

# Workspace directories for temporary files
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMP_DIR = os.path.join(BASE_DIR, "temp_uploads")
STATIC_DIR = os.path.join(BASE_DIR, "static")
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")

# Ensure directories exist
os.makedirs(TEMP_DIR, exist_ok=True)
os.makedirs(STATIC_DIR, exist_ok=True)
os.makedirs(TEMPLATES_DIR, exist_ok=True)

# Mount static folder (for css, js)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# Cleanup helper
def cleanup_files(*filepaths: str):
    for filepath in filepaths:
        if filepath and os.path.exists(filepath):
            try:
                os.remove(filepath)
            except Exception as e:
                print(f"Error removing temp file {filepath}: {e}")

# Cell style copier
def copy_cell_style(src_cell, tgt_cell):
    """Deep copy formatting attributes from src_cell to tgt_cell."""
    if src_cell.has_style:
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.number_format = copy(src_cell.number_format)
        tgt_cell.protection = copy(src_cell.protection)
        tgt_cell.alignment = copy(src_cell.alignment)

# Safe type converter
def clean_and_convert(val, target_cell):
    """Gracefully cast source value into the target cell's expected data type."""
    if val is None:
        return None
    
    # If already native numeric or datetime, return as-is
    if isinstance(val, (int, float, datetime.datetime, datetime.date)):
        return val
    
    val_str = str(val).strip()
    if val_str == "" or val_str.lower() == "nan" or val_str.lower() == "none":
        return None

    # Check target cell number format to determine intent
    num_fmt = target_cell.number_format
    
    # Check if target is a date format
    is_date_fmt = False
    if num_fmt and any(sub in num_fmt.lower() for sub in ["yy", "mm", "dd"]):
        is_date_fmt = True

    if is_date_fmt:
        try:
            # Let pandas parse standard string date styles
            parsed_date = pd.to_datetime(val_str)
            if not pd.isna(parsed_date):
                if parsed_date.time() == datetime.time(0, 0):
                    return parsed_date.date()
                return parsed_date
        except Exception:
            pass # Fall back to other formats or original text

    # Try numeric conversion
    # Strip symbols that represent styling, but might interfere with float/int casting
    clean_num_str = val_str.replace("$", "").replace(",", "").replace("%", "").strip()

    # Attempt casting to float/int
    try:
        if "." in clean_num_str:
            val_num = float(clean_num_str)
        else:
            val_num = int(clean_num_str)
            
        # If the original string had a percentage sign, it's a percentage, so divide by 100
        if "%" in val_str:
            return val_num / 100.0
            
        return val_num
    except ValueError:
        pass

    # Fallback to original string if no format match succeeded
    return val_str


class ColumnRequest(BaseModel):
    file_id: str
    sheet_name: str
    header_row: int = 1

class MergeRequest(BaseModel):
    source_id: str
    target_id: str
    source_sheet: str
    target_sheet: str
    mappings: Dict[str, str]  # target_col -> source_col
    merge_mode: str  # "fill" or "append"
    source_header_row: int = 1
    target_header_row: int = 1

# Route to serve the frontend single page app
@app.get("/", response_class=HTMLResponse)
async def serve_index():
    index_path = os.path.join(TEMPLATES_DIR, "index.html")
    if not os.path.exists(index_path):
        raise HTTPException(status_code=404, detail="Frontend file templates/index.html not found.")
    with open(index_path, "r", encoding="utf-8") as f:
        return f.read()

# Route to upload files and extract sheet names
@app.post("/upload")
async def upload_files(
    file_source: UploadFile = File(...),
    file_target: UploadFile = File(...)
):
    # Validate extensions
    for f in [file_source, file_target]:
        if not f.filename.endswith((".xlsx", ".xlsm")):
            return JSONResponse(
                status_code=400,
                content={"detail": "Only Excel files (.xlsx, .xlsm) are supported."}
            )

    # Save files with a unique prefix
    session_id = uuid.uuid4().hex
    source_filename = f"source_{session_id}_{file_source.filename}"
    target_filename = f"target_{session_id}_{file_target.filename}"
    
    source_path = os.path.join(TEMP_DIR, source_filename)
    target_path = os.path.join(TEMP_DIR, target_filename)

    try:
        with open(source_path, "wb") as buffer:
            shutil.copyfileobj(file_source.file, buffer)
        with open(target_path, "wb") as buffer:
            shutil.copyfileobj(file_target.file, buffer)
    except Exception as e:
        cleanup_files(source_path, target_path)
        raise HTTPException(status_code=500, detail=f"Failed to save uploaded files: {str(e)}")

    # Extract sheets
    try:
        # Load workbooks in read-only mode for sheet retrieval (fast)
        wb_source = load_workbook(source_path, read_only=True)
        source_sheets = wb_source.sheetnames
        wb_source.close()

        wb_target = load_workbook(target_path, read_only=True)
        target_sheets = wb_target.sheetnames
        wb_target.close()
    except Exception as e:
        cleanup_files(source_path, target_path)
        raise HTTPException(status_code=400, detail=f"Failed to read sheet names: {str(e)}")

    return {
        "source_id": source_filename,
        "target_id": target_filename,
        "source_sheets": source_sheets,
        "target_sheets": target_sheets
    }

# Route to extract column headers from a specific sheet
@app.post("/columns")
async def get_columns(req: ColumnRequest):
    file_path = os.path.join(TEMP_DIR, req.file_id)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Uploaded file not found or session expired.")

    try:
        # Load in fast read-only mode
        wb = load_workbook(file_path, read_only=True)
        if req.sheet_name not in wb.sheetnames:
            wb.close()
            raise HTTPException(status_code=400, detail=f"Sheet '{req.sheet_name}' not found in workbook.")
        
        sheet = wb[req.sheet_name]
        
        # Override dimensions to bypass read-only metadata limits
        sheet._max_row = None
        sheet._max_column = None
        
        # Fetch the specified header row values sequentially (supports custom row positioning fast)
        rows_generator = sheet.iter_rows(values_only=True)
        header_row_vals = []
        for r_idx, row in enumerate(rows_generator, start=1):
            if r_idx == req.header_row:
                header_row_vals = row
                break
        
        wb.close()

        # Build clean column headers
        columns = []
        if header_row_vals:
            for idx, cell_val in enumerate(header_row_vals):
                if cell_val is not None:
                    columns.append(str(cell_val).strip())
                else:
                    columns.append(f"Column {idx + 1}")
        
        return {"columns": columns}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read headers: {str(e)}")

# Route to perform mapping and merge the sheets
@app.post("/merge")
async def merge_data(req: MergeRequest, background_tasks: BackgroundTasks):
    source_path = os.path.join(TEMP_DIR, req.source_id)
    target_path = os.path.join(TEMP_DIR, req.target_id)

    if not os.path.exists(source_path) or not os.path.exists(target_path):
        raise HTTPException(status_code=404, detail="Upload files not found. Please upload again.")

    # Create path for output file
    output_filename = f"merged_{uuid.uuid4().hex}.xlsx"
    output_path = os.path.join(TEMP_DIR, output_filename)

    try:
        # 1. Load source values
        wb_source = load_workbook(source_path, data_only=True)
        if req.source_sheet not in wb_source.sheetnames:
            wb_source.close()
            raise HTTPException(status_code=400, detail=f"Source sheet '{req.source_sheet}' not found.")
        
        sheet_source = wb_source[req.source_sheet]
        
        # Read source data
        header_rows_generator = sheet_source.iter_rows(min_row=req.source_header_row, max_row=req.source_header_row, values_only=True)
        try:
            source_headers = next(header_rows_generator)
        except StopIteration:
            source_headers = []
        
        source_headers_cleaned = []
        for idx, h in enumerate(source_headers):
            if h is not None:
                source_headers_cleaned.append(str(h).strip())
            else:
                source_headers_cleaned.append(f"Column {idx + 1}")
        
        # Read source data rows starting after the header row
        rows_iter = sheet_source.iter_rows(min_row=req.source_header_row + 1, values_only=True)
        source_data = list(rows_iter)
        wb_source.close()

        # Mapping source headers to tuple indices
        source_cols_map = {name: idx for idx, name in enumerate(source_headers_cleaned)}

        # 2. Load target workbook in write mode to preserve formatting
        wb_target = load_workbook(target_path)
        if req.target_sheet not in wb_target.sheetnames:
            wb_target.close()
            raise HTTPException(status_code=400, detail=f"Target sheet '{req.target_sheet}' not found.")
        
        sheet_target = wb_target[req.target_sheet]

        # Extract target headers and columns index from specified target_header_row
        target_cols = {}
        for col in range(1, sheet_target.max_column + 1):
            val = sheet_target.cell(row=req.target_header_row, column=col).value
            val_str = str(val).strip() if val is not None else f"Column {col}"
            target_cols[val_str] = col

        # Calculate template row parameters
        original_max_row = sheet_target.max_row
        
        # Find the actual last data row (by looking for any non-empty cell starting from target_header_row + 1)
        last_data_row = req.target_header_row
        for r in range(req.target_header_row + 1, original_max_row + 1):
            has_value = False
            for c in range(1, sheet_target.max_column + 1):
                if sheet_target.cell(row=r, column=c).value is not None:
                    has_value = True
                    break
            if has_value:
                last_data_row = r

        # Set start row
        if req.merge_mode == "append":
            start_row = last_data_row + 1
        else: # "fill"
            start_row = req.target_header_row + 1

        # Style source map for extension copying (reference row target_header_row + 1 if exists, otherwise target_header_row)
        style_ref_row = req.target_header_row + 1 if original_max_row >= req.target_header_row + 1 else req.target_header_row
        col_styles = {}
        for col_idx in target_cols.values():
            col_styles[col_idx] = sheet_target.cell(row=style_ref_row, column=col_idx)

        # 3. Perform Merge
        for i, src_row in enumerate(source_data):
            target_row_idx = start_row + i
            
            # Check if any mapped source fields have data for this row (skipping fully empty tailing rows)
            has_data = False
            for src_col, tgt_col in req.mappings.items():
                if src_col in source_cols_map:
                    val = src_row[source_cols_map[src_col]]
                    if val is not None and str(val).strip() != "":
                        has_data = True
                        break
            
            if not has_data:
                continue  # Skip empty rows from the source sheet

            # Write row columns
            for src_col_name, tgt_col_name in req.mappings.items():
                if tgt_col_name not in target_cols or src_col_name not in source_cols_map:
                    continue
                
                tgt_col_idx = target_cols[tgt_col_name]
                src_col_idx = source_cols_map[src_col_name]

                
                src_val = src_row[src_col_idx]
                target_cell = sheet_target.cell(row=target_row_idx, column=tgt_col_idx)
                
                # Clone style if row extends past original template length
                if target_row_idx > original_max_row:
                    ref_cell = col_styles.get(tgt_col_idx)
                    if ref_cell:
                        copy_cell_style(ref_cell, target_cell)
                
                # Graceful typecast conversion
                target_cell.value = clean_and_convert(src_val, target_cell)

        # Save merged file
        wb_target.save(output_path)
        wb_target.close()

    except Exception as e:
        cleanup_files(output_path)
        raise HTTPException(status_code=500, detail=f"Excel merging failed: {str(e)}")

    # Schedule cleanup of temp files in the background
    background_tasks.add_task(cleanup_files, source_path, target_path, output_path)

    # Return the file stream
    original_target_name = req.target_id.split(f"target_{req.target_id.split('_')[1]}_")[-1]
    download_name = f"FlexiMapper_{original_target_name}"
    
    return FileResponse(
        path=output_path,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
