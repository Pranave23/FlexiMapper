import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import the merge functions from app
from app import clean_and_convert, copy_cell_style

def create_sample_source():
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    # Headers
    headers = ["ID", "First Name", "Last Name", "Salary", "Joining Date"]
    ws.append(headers)
    
    # Data
    data = [
        [101, "Alice", "Smith", "$85,000", "2023-01-15"],
        [102, "Bob", "Jones", "92,500.50", "2022-06-01"],
        [103, "Charlie", "Brown", "78,000%", "2024-03-10"], # test messy percentage string
        [104, "Diana", "Prince", "110000", "2021-11-20"]
    ]
    for row in data:
        ws.append(row)
        
    source_path = "test_source.xlsx"
    wb.save(source_path)
    wb.close()
    return source_path

def create_sample_target():
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Setup some headers
    headers = ["ID", "First_Name", "Last_Name", "Income", "Start_Date", "Unmapped_Col"]
    ws.append(headers)
    
    # Format Headers (Bold font, Blue background fill, Thin borders)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
    # Setup Row 2 with sample formatting to act as the template base styles
    data_font = Font(name="Calibri", size=11, italic=True)
    ws.cell(row=2, column=1).font = data_font # ID
    ws.cell(row=2, column=2).font = data_font # First_Name
    ws.cell(row=2, column=3).font = data_font # Last_Name
    
    # Income column formatted as currency
    income_cell = ws.cell(row=2, column=4)
    income_cell.font = data_font
    income_cell.number_format = '"$"#,##0.00'
    
    # Start_Date column formatted as Date
    date_cell = ws.cell(row=2, column=5)
    date_cell.font = data_font
    date_cell.number_format = 'yyyy-mm-dd'
    
    # Unmapped column
    unmapped_cell = ws.cell(row=2, column=6)
    unmapped_cell.value = "TEMPLATE_DEFAULT_VAL"
    unmapped_cell.font = data_font
    
    target_path = "test_target.xlsx"
    wb.save(target_path)
    wb.close()
    return target_path

def run_test_merge(source_path, target_path, mode="fill"):
    # Mocking the merge logic programmatically
    from openpyxl import load_workbook
    
    wb_source = load_workbook(source_path, data_only=True)
    sheet_source = wb_source["Employees"]
    
    # Read source
    rows_iter = sheet_source.iter_rows(values_only=True)
    source_headers = next(rows_iter)
    source_headers_cleaned = [str(h).strip() if h is not None else f"Column {i+1}" for i, h in enumerate(source_headers)]
    source_data = list(rows_iter)
    wb_source.close()
    
    source_cols_map = {name: idx for idx, name in enumerate(source_headers_cleaned)}

    wb_target = load_workbook(target_path)
    sheet_target = wb_target["Template"]
    
    # Target columns map
    target_cols = {}
    for col in range(1, sheet_target.max_column + 1):
        val = sheet_target.cell(row=1, column=col).value
        val_str = str(val).strip() if val is not None else f"Column {col}"
        target_cols[val_str] = col
        
    mappings = {
        "ID": "ID",
        "First Name": "First_Name",
        "Last Name": "Last_Name",
        "Salary": "Income",
        "Joining Date": "Start_Date"
    }
    
    original_max_row = sheet_target.max_row
    
    # Determine last data row
    last_data_row = 1
    for r in range(2, original_max_row + 1):
        has_value = False
        for c in range(1, sheet_target.max_column + 1):
            if sheet_target.cell(row=r, column=c).value is not None:
                has_value = True
                break
        if has_value:
            last_data_row = r
            
    if mode == "append":
        start_row = last_data_row + 1
    else:
        start_row = 2
        
    # Styles referencing row 2
    style_ref_row = 2 if original_max_row >= 2 else 1
    col_styles = {col_idx: sheet_target.cell(row=style_ref_row, column=col_idx) for col_idx in target_cols.values()}
    
    for i, src_row in enumerate(source_data):
        target_row_idx = start_row + i
        
        for src_col_name, tgt_col_name in mappings.items():
            if tgt_col_name not in target_cols or src_col_name not in source_cols_map:
                continue
            
            tgt_col_idx = target_cols[tgt_col_name]
            src_col_idx = source_cols_map[src_col_name]

            src_val = src_row[src_col_idx]
            
            target_cell = sheet_target.cell(row=target_row_idx, column=tgt_col_idx)
            
            # Style clone if extending
            if target_row_idx > original_max_row:
                ref_cell = col_styles.get(tgt_col_idx)
                if ref_cell:
                    copy_cell_style(ref_cell, target_cell)
                    
            # Type cast conversion
            target_cell.value = clean_and_convert(src_val, target_cell)
            
    output_path = "test_merged_output.xlsx"
    wb_target.save(output_path)
    wb_target.close()
    return output_path

def verify_output(output_path):
    from openpyxl import load_workbook
    wb = load_workbook(output_path)
    ws = wb["Template"]
    
    print("\n--- MERGED WORKBOOK VERIFICATION ---")
    print(f"Max Row: {ws.max_row}")
    print(f"Max Column: {ws.max_column}")
    
    # Assert row count
    # Row 1: Headers
    # Rows 2, 3, 4, 5: Merged Data (4 rows of data)
    assert ws.max_row == 5, f"Expected 5 rows, got {ws.max_row}"
    
    # Check Row 2 data (Alice)
    assert ws.cell(row=2, column=1).value == 101
    assert ws.cell(row=2, column=2).value == "Alice"
    assert ws.cell(row=2, column=3).value == "Smith"
    assert ws.cell(row=2, column=4).value == 85000.0
    assert isinstance(ws.cell(row=2, column=5).value, (datetime.date, datetime.datetime))
    assert ws.cell(row=2, column=6).value == "TEMPLATE_DEFAULT_VAL" # untouched template cell value
    
    # Check Row 4 data (Charlie)
    assert ws.cell(row=4, column=1).value == 103
    assert ws.cell(row=4, column=2).value == "Charlie"
    assert ws.cell(row=4, column=4).value == 780.0  # 78,000% parsed to float
    
    # Verify Styles propagation for overflow rows (Row 5 - Diana)
    diana_income_cell = ws.cell(row=5, column=4)
    assert diana_income_cell.value == 110000
    assert diana_income_cell.font.italic is True, "Font style failed to clone"
    assert diana_income_cell.number_format == '"$"#,##0.00', "Number format failed to clone"
    
    diana_date_cell = ws.cell(row=5, column=5)
    assert diana_date_cell.number_format == 'yyyy-mm-dd', "Date number format failed to clone"
    
    # Verify unmapped cells in new rows remain empty
    diana_unmapped_cell = ws.cell(row=5, column=6)
    assert diana_unmapped_cell.value is None
    
    print("\n[SUCCESS] All assertions passed!")
    print("- Excel file values were mapped properly.")
    print("- Number conversions (e.g. '$85,000' -> 85000) succeeded.")
    print("- Extended rows (rows 3-5) correctly cloned font styles and number formats from the template row.")
    wb.close()

if __name__ == "__main__":
    print("Generating sample spreadsheets...")
    src = create_sample_source()
    tgt = create_sample_target()
    
    print("Executing merge engine programmatically...")
    out = run_test_merge(src, tgt, mode="fill")
    
    print("Verifying merged output specifications...")
    verify_output(out)
    
    # Clean up test files
    for f in [src, tgt, out]:
        if os.path.exists(f):
            os.remove(f)
            print(f"Cleaned up {f}")
