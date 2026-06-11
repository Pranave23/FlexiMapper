import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
import datetime
from copy import copy
import re


# Custom Page Config
st.set_page_config(
    page_title="FlexiMapper — Intelligent Column Mapping",
    page_icon="🗺️",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom premium styling via CSS injection
st.markdown("""
    <style>
    /* Styling headers and branding */
    .main-title {
        font-family: 'Plus Jakarta Sans', sans-serif;
        font-weight: 800;
        background: linear-gradient(90deg, #6366f1 0%, #a78bfa 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 3rem;
        margin-bottom: 0.2rem;
    }
    .subtitle {
        font-size: 1.1rem;
        color: #9ca3af;
        margin-bottom: 2rem;
    }
    /* Section dividers */
    .section-header {
        font-weight: 700;
        border-bottom: 2px solid #374151;
        padding-bottom: 0.5rem;
        margin-top: 1.5rem;
        margin-bottom: 1rem;
        color: #f3f4f6;
    }
    /* Buttons */
    div.stButton > button {
        background: linear-gradient(90deg, #4f46e5 0%, #6366f1 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.6rem 1.5rem;
        font-weight: bold;
        transition: all 0.3s ease;
    }
    div.stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 12px rgba(99, 102, 241, 0.3);
    }
    </style>
""", unsafe_allow_html=True)

# Branding Header
st.markdown("<h1 class='main-title'>FlexiMapper</h1>", unsafe_allow_html=True)
st.markdown("<p class='subtitle'>Map Excel columns visually and merge data while preserving cell formatting.</p>", unsafe_allow_html=True)

# Helper functions for styles cloning & conversion
def copy_cell_style(src_cell, tgt_cell):
    """Deep copy formatting attributes from src_cell to tgt_cell."""
    if src_cell.has_style:
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.number_format = copy(src_cell.number_format)
        tgt_cell.protection = copy(src_cell.protection)
        tgt_cell.alignment = copy(src_cell.alignment)

def clean_and_convert(val, target_cell):
    """Gracefully cast source value into the target cell's expected data type."""
    if val is None:
        return None
    if isinstance(val, (int, float, datetime.datetime, datetime.date)):
        return val
    
    val_str = str(val).strip()
    if val_str == "" or val_str.lower() == "nan" or val_str.lower() == "none":
        return None

    num_fmt = target_cell.number_format
    
    # Check if target expects date
    is_date_fmt = False
    if num_fmt and any(sub in num_fmt.lower() for sub in ["yy", "mm", "dd"]):
        is_date_fmt = True

    if is_date_fmt:
        try:
            parsed_date = pd.to_datetime(val_str)
            if not pd.isna(parsed_date):
                if parsed_date.time() == datetime.time(0, 0):
                    return parsed_date.date()
                return parsed_date
        except Exception:
            pass

    # Clean numeric format (currency, percentages, commas)
    clean_num_str = val_str.replace("$", "").replace(",", "").replace("%", "").strip()

    try:
        if "." in clean_num_str:
            val_num = float(clean_num_str)
        else:
            val_num = int(clean_num_str)
            
        if "%" in val_str:
            return val_num / 100.0
            
        return val_num
    except ValueError:
        pass

    return val_str

# String similarity utility for smart match
def clean_string(s):
    if isinstance(s, str):
        return re.sub(r'[^a-z0-9]', '', s.lower())
    return str(s).lower()

def get_similarity_score(s1, s2):
    s1_clean = "".join(c for c in str(s1).lower() if c.isalnum())
    s2_clean = "".join(c for c in str(s2).lower() if c.isalnum())
    if s1_clean == s2_clean:
        return 1.0
    
    len1, len2 = len(s1_clean), len(s2_clean)
    max_len = max(len1, len2)
    if max_len == 0:
        return 1.0
        
    # Levenshtein distance
    dp = [[0] * (len2 + 1) for _ in range(len1 + 1)]
    for i in range(len1 + 1):
        dp[i][0] = i
    for j in range(len2 + 1):
        dp[0][j] = j
        
    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            if s1_clean[i-1] == s2_clean[j-1]:
                dp[i][j] = dp[i-1][j-1]
            else:
                dp[i][j] = min(dp[i-1][j-1] + 1, dp[i][j-1] + 1, dp[i-1][j] + 1)
                
    distance = dp[len1][len2]
    score = (max_len - distance) / max_len
    
    # Give boost for containment
    if s1_clean in s2_clean or s2_clean in s1_clean:
        score *= 1.15
    return min(score, 1.0)

# STEP 1: File Uploaders
st.markdown("<h3 class='section-header'>1. Upload Spreadsheets</h3>", unsafe_allow_html=True)
col_file1, col_file2 = st.columns(2)

with col_file1:
    source_file = st.file_uploader("Source Sheet (Sheet 1) — Data Origin", type=["xlsx", "xlsm"])
with col_file2:
    target_file = st.file_uploader("Target Sheet / Template (Sheet 2) — Data Destination", type=["xlsx", "xlsm"])

# If both files are uploaded, parse sheets
if source_file and target_file:
    try:
        # Load sheets lists using openpyxl (read_only is fast)
        wb_src = load_workbook(source_file, read_only=True)
        src_sheets = wb_src.sheetnames
        wb_src.close()
        source_file.seek(0) # reset stream

        wb_tgt = load_workbook(target_file, read_only=True)
        tgt_sheets = wb_tgt.sheetnames
        wb_tgt.close()
        target_file.seek(0) # reset stream
        
        # Display sheet selectors
        col_sh1, col_sh2 = st.columns(2)
        with col_sh1:
            src_sheet_selected = st.selectbox("Select Source Sheet Tab:", src_sheets)
        with col_sh2:
            tgt_sheet_selected = st.selectbox("Select Target Sheet Tab:", tgt_sheets)

        # Read column headers
        wb_src = load_workbook(source_file, read_only=True)
        ws_src = wb_src[src_sheet_selected]
        src_row_gen = ws_src.iter_rows(values_only=True)
        try:
            src_first_row = next(src_row_gen)
        except StopIteration:
            src_first_row = []
        wb_src.close()
        source_file.seek(0)

        wb_tgt = load_workbook(target_file, read_only=True)
        ws_tgt = wb_tgt[tgt_sheet_selected]
        tgt_row_gen = ws_tgt.iter_rows(values_only=True)
        try:
            tgt_first_row = next(tgt_row_gen)
        except StopIteration:
            tgt_first_row = []
        wb_tgt.close()
        target_file.seek(0)

        # Clean source headers
        source_columns = []
        if src_first_row:
            for idx, c in enumerate(src_first_row):
                source_columns.append(str(c).strip() if c is not None else f"Column {idx+1}")

        # Clean target headers
        target_columns = []
        if tgt_first_row:
            for idx, c in enumerate(tgt_first_row):
                target_columns.append(str(c).strip() if c is not None else f"Column {idx+1}")

        # STEP 2: Mappings Panel
        st.markdown("<h3 class='section-header'>2. Align Columns</h3>", unsafe_allow_html=True)
        
        # Smart Match Button
        if st.button("Smart Match Columns"):
            st.session_state.smart_matched = True
            st.session_state.mappings = {}
            for src_col in source_columns:
                best_match = ""
                highest_score = 0.0
                for tgt_col in target_columns:
                    score = get_similarity_score(src_col, tgt_col)
                    if score > highest_score:
                        highest_score = score
                        best_match = tgt_col
                if highest_score >= 0.6:
                    st.session_state.mappings[src_col] = best_match
                else:
                    st.session_state.mappings[src_col] = "— Skip —"
        
        if 'mappings' not in st.session_state:
            st.session_state.mappings = {}

        # Render mapping grid (Source LHS, Target RHS)
        mappings = {}
        for src_col in source_columns:
            col_lhs, col_arrow, col_rhs = st.columns([2, 1, 3])
            with col_lhs:
                st.write("") # spacing
                st.write(f"**{src_col}**")
            with col_arrow:
                st.write("")
                st.write("➡️")
            with col_rhs:
                # Find default index based on session state mappings
                default_val = st.session_state.mappings.get(src_col, "— Skip —")
                options_list = ["— Skip —"] + target_columns
                default_idx = options_list.index(default_val) if default_val in options_list else 0
                
                selected_target = st.selectbox(
                    f"Select Target Column for {src_col}",
                    options_list,
                    index=default_idx,
                    key=f"select_{src_col}",
                    label_visibility="collapsed"
                )
                if selected_target != "— Skip —":
                    mappings[src_col] = selected_target

        # STEP 3: Options & Execution
        st.markdown("<h3 class='section-header'>3. Merge Options</h3>", unsafe_allow_html=True)
        
        merge_mode = st.radio(
            "Select Merge Strategy:",
            options=["Fill / Overwrite Columns (Writes starting at Row 2)", 
                     "Append to Existing Data (Writes after the last data row)"],
            index=0
        )
        strategy = "fill" if "Fill" in merge_mode else "append"

        st.write("")
        if st.button("Process & Merge Data"):
            if not mappings:
                st.warning("Please map at least one column before merging.")
            else:
                with st.spinner("Processing spreadsheets..."):
                    try:
                        # 1. Load source rows
                        wb_source = load_workbook(source_file, data_only=True)
                        ws_source = wb_source[src_sheet_selected]
                        
                        src_rows_iter = ws_source.iter_rows(values_only=True)
                        next(src_rows_iter) # skip headers
                        source_data = list(src_rows_iter)
                        wb_source.close()
                        source_file.seek(0)
                        
                        source_cols_map = {name: idx for idx, name in enumerate(source_columns)}

                        # 2. Load target workbook in write-mode (keep file format/styles)
                        wb_target = load_workbook(target_file)
                        ws_target = wb_target[tgt_sheet_selected]

                        target_cols_idx = {}
                        for col in range(1, ws_target.max_column + 1):
                            val = ws_target.cell(row=1, column=col).value
                            val_str = str(val).strip() if val is not None else f"Column {col}"
                            target_cols_idx[val_str] = col

                        original_max_row = ws_target.max_row

                        # Calculate active last data row in Target
                        last_data_row = 1
                        for r in range(2, original_max_row + 1):
                            has_val = False
                            for c in range(1, ws_target.max_column + 1):
                                if ws_target.cell(row=r, column=c).value is not None:
                                    has_val = True
                                    break
                            if has_val:
                                last_data_row = r

                        # Starting row index
                        start_row = last_data_row + 1 if strategy == "append" else 2

                        # Store styling references from row 2 (or row 1 if single row template)
                        style_ref_row = 2 if original_max_row >= 2 else 1
                        col_styles = {col_idx: ws_target.cell(row=style_ref_row, column=col_idx) for col_idx in target_cols_idx.values()}

                        # 3. Paste data row by row
                        for i, src_row in enumerate(source_data):
                            target_row_idx = start_row + i

                            # Skip empty tail rows
                            has_data = False
                            for src_col, tgt_col in mappings.items():
                                if src_col in source_cols_map:
                                    val = src_row[source_cols_map[src_col]]
                                    if val is not None and str(val).strip() != "":
                                        has_data = True
                                        break
                            if not has_data:
                                continue

                            # Copy values
                            for src_col_name, tgt_col_name in mappings.items():
                                if tgt_col_name not in target_cols_idx or src_col_name not in source_cols_map:
                                    continue

                                tgt_col_idx = target_cols_idx[tgt_col_name]
                                src_col_idx = source_cols_map[src_col_name]
                                src_val = src_row[src_col_idx]

                                target_cell = ws_target.cell(row=target_row_idx, column=tgt_col_idx)

                                # Clone styles if row extends past original template length
                                if target_row_idx > original_max_row:
                                    ref_cell = col_styles.get(tgt_col_idx)
                                    if ref_cell:
                                        copy_cell_style(ref_cell, target_cell)

                                # Convert types
                                target_cell.value = clean_and_convert(src_val, target_cell)

                        # Save to a memory buffer
                        output_buffer = io.BytesIO()
                        wb_target.save(output_buffer)
                        output_buffer.seek(0)
                        wb_target.close()
                        target_file.seek(0)

                        st.success("Successfully Merged Sheet Data!")
                        
                        # Download Button
                        st.download_button(
                            label="Download Merged Excel File",
                            data=output_buffer,
                            file_name=f"FlexiMapper_{target_file.name}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

                    except Exception as e:
                        st.error(f"Error during merge: {str(e)}")
                        
    except Exception as e:
        st.error(f"Failed to scan spreadsheets: {str(e)}")
else:
    st.info("Please upload both a Source file and a Target template file to begin column mapping.")
